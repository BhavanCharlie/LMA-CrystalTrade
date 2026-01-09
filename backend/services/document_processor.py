import os
import PyPDF2
import pdfplumber
from docx import Document as DocxDocument
import openpyxl
from typing import Dict, Optional
import pytesseract
from PIL import Image
import io


class DocumentProcessor:
    def __init__(self):
        self.supported_types = {
            ".pdf": self._process_pdf,
            ".doc": self._process_docx,
            ".docx": self._process_docx,
            ".xls": self._process_xlsx,
            ".xlsx": self._process_xlsx,
        }

    async def process_document(self, file_path: str, filename: str) -> Dict:
        """Process document and return metadata"""
        file_ext = os.path.splitext(filename)[1].lower()
        
        if file_ext not in self.supported_types:
            raise ValueError(f"Unsupported file type: {file_ext}")
        
        processor = self.supported_types[file_ext]
        doc_type = await self._classify_document(file_path, file_ext)
        
        return {
            "type": doc_type,
            "extension": file_ext,
            "processed": True,
        }

    async def extract_text(self, file_path: str) -> str:
        """Extract text from document"""
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext not in self.supported_types:
            raise ValueError(f"Unsupported file type: {file_ext}")
        
        processor = self.supported_types[file_ext]
        return await processor(file_path)

    async def _process_pdf(self, file_path: str) -> str:
        """Extract text from PDF"""
        text = ""
        try:
            # Try pdfplumber first (better for structured content)
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception:
            # Fallback to PyPDF2
            try:
                with open(file_path, "rb") as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
            except Exception:
                # Try OCR if text extraction fails
                text = await self._ocr_pdf(file_path)
        
        return text

    async def _process_docx(self, file_path: str) -> str:
        """Extract text from Word document"""
        doc = DocxDocument(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])

    async def _process_xlsx(self, file_path: str) -> str:
        """Extract text from Excel file"""
        workbook = openpyxl.load_workbook(file_path)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}\n")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                text_parts.append(row_text)
            text_parts.append("\n")
        
        return "\n".join(text_parts)

    async def _ocr_pdf(self, file_path: str) -> str:
        """OCR for PDF using pytesseract"""
        # This is a placeholder - would need proper PDF to image conversion
        return "OCR text extraction not fully implemented"

    async def _classify_document(self, file_path: str, file_ext: str) -> str:
        """Classify document type based on content"""
        # Simple classification based on filename and content
        filename_lower = os.path.basename(file_path).lower()
        
        if "credit" in filename_lower or "agreement" in filename_lower:
            return "credit_agreement"
        elif "amendment" in filename_lower:
            return "amendment"
        elif "financial" in filename_lower or "statement" in filename_lower:
            return "financial_statement"
        elif "covenant" in filename_lower:
            return "covenant_compliance"
        else:
            return "other"


